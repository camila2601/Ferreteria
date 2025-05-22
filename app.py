from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response, send_file
from pymongo import MongoClient
from bson.objectid import ObjectId
from datetime import datetime
import pandas as pd
from fpdf import FPDF
import io
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import pdfkit
import openpyxl
from io import BytesIO
import os
from functools import wraps
import pymongo
from pymongo import MongoClient, DESCENDING
from bson.objectid import ObjectId
from urllib.parse import quote_plus

# ==============================================
# CONFIGURACIÓN INICIAL Y CONEXIÓN A MONGODB
# ==============================================
app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Contraseña y conexión
password = "JJvxjpLqafSuRNHX"
escaped_password = quote_plus(password)

username = "myAtlasDBUser"
cluster = "ferreteria.hmvlsvu.mongodb.net"
connection_string = f"mongodb+srv://{username}:{escaped_password}@{cluster}/?retryWrites=true&w=majority&appName=ferreteria"




try:
    client = pymongo.MongoClient(connection_string)
    db = client["ferreteria"]
    db.command('ping')
    print("\n" + "="*50)
    print("\u2705 CONEXIÓN EXITOSA A MONGODB ATLAS")
    print(f"📊 Base de datos: {db.name}")
    print(f"📦 Colección 'users': {db.users.count_documents({})} registros")
    print("="*50 + "\n")
except Exception as e:
    print("\n" + "="*50)
    print("❌ ERROR DE CONEXIÓN")
    print(f"Detalle: {str(e)}")
    print("="*50 + "\n")
    raise e

# ==============================================
# FUNCIONES AUXILIARES
# ==============================================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor inicie sesión primero', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# ==============================================
# RUTAS DE AUTENTICACIÓN
# ==============================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            flash('Por favor complete todos los campos', 'danger')
            return redirect(url_for('login'))

        user = db.users.find_one({'username': username})
        if user and check_password_hash(user['password'], password):
            session['user_id'] = str(user['_id'])
            session['username'] = user['username']
            session['role'] = user.get('role', 'user')

            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))

        flash('Usuario o contraseña incorrectos', 'danger')

    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        errors = []
        if not all([username, password, confirm_password]):
            errors.append('Todos los campos son requeridos')
        if password != confirm_password:
            errors.append('Las contraseñas no coinciden')
        if len(password) < 6:
            errors.append('La contraseña debe tener al menos 6 caracteres')
        if db.users.find_one({'username': username}):
            errors.append('El nombre de usuario ya está en uso')

        if errors:
            for error in errors:
                flash(error, 'danger')
            return redirect(url_for('register'))

        try:
            db.users.insert_one({
                'username': username,
                'password': generate_password_hash(password),
                'role': 'user',
                'created_at': datetime.utcnow(),
                'last_login': None
            })
            flash('¡Registro exitoso! Por favor inicie sesión', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            flash(f'Error al registrar: {str(e)}', 'danger')
            return redirect(url_for('register'))

    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    session.clear()
    flash('Sesión cerrada correctamente', 'info')
    return redirect(url_for('login'))

# ==============================================
# RUTAS PRINCIPALES
# ==============================================
@app.route('/')
def home():
    return redirect(url_for('dashboard') if 'user_id' in session else 'login')

@app.route('/dashboard')
@login_required
def dashboard():
    total_productos = db.productos.count_documents({})
    total_clientes = db.clientes.count_documents({})
    hoy = datetime.combine(datetime.today(), datetime.min.time())
    ventas_hoy = db.ventas.count_documents({'fecha': {'$gte': hoy}})
    productos_bajo_stock = list(db.productos.find({
        '$expr': {'$lt': ['$stock', '$stock_minimo']}
    }).limit(5))
    ultimas_ventas = list(db.ventas.find().sort('fecha', DESCENDING).limit(5))
    return render_template('dashboard.html',
                         total_productos=total_productos,
                         total_clientes=total_clientes,
                         ventas_hoy=ventas_hoy,
                         productos_bajo_stock=productos_bajo_stock,
                         ultimas_ventas=ultimas_ventas)

# ---------- CRUD PRODUCTOS ----------
@app.route('/productos')
def mostrar_productos():
    all_products = list(db.productos.find())
    return render_template('products.html', productos=all_products)

@app.route('/agregar_producto', methods=['POST'])
def agregar_producto():
    producto = {
        "nombre": request.form["nombre"],
        "codigo_barras": request.form.get("codigo_barras", ""),
        "precio": float(request.form["precio"]),
        "costo": float(request.form["costo"]),
        "stock": int(request.form["stock"]),
        "stock_minimo": int(request.form["stock_minimo"]),
        "categoria": request.form["categoria"],
        "proveedor": request.form.get("proveedor", ""),
        "descripcion": request.form.get("descripcion", "")
    }
    db.productos.insert_one(producto)
    return redirect(url_for('mostrar_productos'))

@app.route('/eliminar_producto/<id>')
def eliminar_producto(id):
    db.productos.delete_one({"_id": ObjectId(id)})
    return redirect(url_for('mostrar_productos'))

@app.route('/nuevo_producto')
def nuevo_producto():
    categorias = list(db.categorias.find()) if 'categorias' in db.list_collection_names() else []
    proveedores = list(db.proveedores.find()) if 'proveedores' in db.list_collection_names() else []
    return render_template('add_product.html', categorias=categorias, proveedores=proveedores)

# ---------- CRUD CLIENTES ----------
@app.route('/clientes')
def mostrar_clientes():
    all_clients = list(db.clientes.find())
    return render_template('clients.html', clientes=all_clients)

@app.route('/agregar_cliente', methods=['POST'])
def agregar_cliente():
    cliente = {
        "nombre": request.form["nombre"],
        "correo": request.form["correo"],
        "telefono": request.form["telefono"]
    }
    db.clientes.insert_one(cliente)
    return redirect(url_for('mostrar_clientes'))

@app.route('/eliminar_cliente/<id>')
def eliminar_cliente(id):
    db.clientes.delete_one({"_id": ObjectId(id)})
    return redirect(url_for('mostrar_clientes'))

# ---------- VENTAS Y DEVOLUCIONES ----------
@app.route('/ventas')
def mostrar_ventas():
    all_ventas = list(db.ventas.find())
    all_productos = list(db.productos.find())
    all_clientes = list(db.clientes.find())
    return render_template('sales.html', ventas=all_ventas, productos=all_productos, clientes=all_clientes)

@app.route('/realizar_venta', methods=['POST'])
def realizar_venta():
    id_producto = request.form["producto"]
    id_cliente = request.form["cliente"]
    cantidad = int(request.form["cantidad"])

    producto = db.productos.find_one({"_id": ObjectId(id_producto)})
    if producto and producto["stock"] >= cantidad:
        nueva_stock = producto["stock"] - cantidad
        db.productos.update_one({"_id": ObjectId(id_producto)}, {"$set": {"stock": nueva_stock}})
        venta = {
            "id_producto": id_producto,
            "id_cliente": id_cliente,
            "cantidad": cantidad,
            "total": cantidad * producto["precio"],
            "fecha": datetime.now(),
            "tipo": "venta"
        }
        db.ventas.insert_one(venta)
    return redirect(url_for('mostrar_ventas'))

@app.route('/devolucion/<id>')
def devolucion(id):
    venta = db.ventas.find_one({"_id": ObjectId(id)})
    if venta and venta["tipo"] == "venta":
        producto = db.productos.find_one({"_id": ObjectId(venta["id_producto"])})
        nueva_stock = producto["stock"] + venta["cantidad"]
        db.productos.update_one({"_id": ObjectId(venta["id_producto"])}, {"$set": {"stock": nueva_stock}})
        devolucion = venta.copy()
        devolucion["_id"] = ObjectId()
        devolucion["tipo"] = "devolucion"
        devolucion["fecha"] = datetime.now()
        db.ventas.insert_one(devolucion)
    return redirect(url_for('mostrar_ventas'))

# ---------- REPORTES ----------
@app.route('/reportes')
def reportes():
    return render_template('reports.html')

@app.route('/reporte_excel')
def reporte_excel():
    data = list(db.ventas.find())
    for d in data:
        d["_id"] = str(d["_id"])
        d["fecha"] = d["fecha"].strftime('%Y-%m-%d %H:%M:%S')
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Ventas', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Ventas']
        # Formato de encabezado
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'middle',
            'align': 'center',
            'fg_color': '#FFB366',
            'border': 1
        })
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 18)
        # Bordes para toda la tabla
        table_format = workbook.add_format({'border': 1})
        worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': table_format})
    output.seek(0)
    return send_file(output, download_name='reporte_ventas.xlsx', as_attachment=True)

@app.route('/reporte_pdf')
def reporte_pdf():
    data = list(db.ventas.find())
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, txt="Reporte de Ventas", ln=True, align='C')
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(230, 230, 230)
    # Encabezados de tabla
    pdf.cell(45, 10, 'ProductoID', 1, 0, 'C', 1)
    pdf.cell(45, 10, 'ClienteID', 1, 0, 'C', 1)
    pdf.cell(25, 10, 'Cantidad', 1, 0, 'C', 1)
    pdf.cell(30, 10, 'Total', 1, 0, 'C', 1)
    pdf.cell(35, 10, 'Fecha', 1, 0, 'C', 1)
    pdf.cell(25, 10, 'Tipo', 1, 1, 'C', 1)
    pdf.set_font("Arial", '', 11)
    row_height = 9
    for v in data:
        pdf.cell(45, row_height, str(v.get('id_producto','')), 1, 0, 'C')
        pdf.cell(45, row_height, str(v.get('id_cliente','')), 1, 0, 'C')
        pdf.cell(25, row_height, str(v.get('cantidad','')), 1, 0, 'C')
        pdf.cell(30, row_height, f"${v.get('total',0):.2f}", 1, 0, 'C')
        fecha = v.get('fecha')
        if fecha:
            fecha_str = fecha.strftime('%Y-%m-%d')
        else:
            fecha_str = ''
        pdf.cell(35, row_height, fecha_str, 1, 0, 'C')
        pdf.cell(25, row_height, str(v.get('tipo','')), 1, 1, 'C')
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    output = io.BytesIO(pdf_bytes)
    output.seek(0)
    return send_file(output, download_name='reporte_ventas.pdf', as_attachment=True)

@app.route('/reporte_inventario_pdf')
def reporte_inventario_pdf():
    productos = list(db.productos.find())
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, txt="Reporte de Inventario", ln=True, align='C')
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(50, 10, 'Nombre', 1, 0, 'C', 1)
    pdf.cell(35, 10, 'Código', 1, 0, 'C', 1)
    pdf.cell(25, 10, 'Stock', 1, 0, 'C', 1)
    pdf.cell(30, 10, 'Stock Min.', 1, 0, 'C', 1)
    pdf.cell(30, 10, 'Precio', 1, 0, 'C', 1)
    pdf.cell(40, 10, 'Categoría', 1, 1, 'C', 1)
    pdf.set_font("Arial", '', 11)
    for p in productos:
        pdf.cell(50, 8, str(p.get('nombre','')), 1)
        pdf.cell(35, 8, str(p.get('codigo_barras','')), 1)
        pdf.cell(25, 8, str(p.get('stock',0)), 1)
        pdf.cell(30, 8, str(p.get('stock_minimo',0)), 1)
        pdf.cell(30, 8, f"${p.get('precio',0):.2f}", 1)
        pdf.cell(40, 8, str(p.get('categoria','')), 1, 1)
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    output = io.BytesIO(pdf_bytes)
    output.seek(0)
    return send_file(output, download_name='reporte_inventario.pdf', as_attachment=True)

@app.route('/reporte_inventario_excel')
def reporte_inventario_excel():
    productos = list(db.productos.find())
    for p in productos:
        p['_id'] = str(p['_id'])
    df = pd.DataFrame(productos)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Inventario']
        # Formato de encabezado
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'middle',
            'align': 'center',
            'fg_color': '#B6D7A8',
            'border': 1
        })
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 18)
        # Bordes para toda la tabla
        table_format = workbook.add_format({'border': 1})
        worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': table_format})
    output.seek(0)
    return send_file(output, download_name='reporte_inventario.xlsx', as_attachment=True)

# ==============================================
# RUTA DE ESTADO DE LA BASE DE DATOS
# ==============================================
@app.route('/db-status')
def db_status():
    try:
        return {
            'status': 'success',
            'db': db.name,
            'collections': db.list_collection_names(),
            'users_count': db.users.count_documents({}),
            'app': 'Ferretería App'
        }
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

# ==============================================
# INICIO DE LA APLICACIÓN
# ==============================================
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
